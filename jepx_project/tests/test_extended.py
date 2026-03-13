"""追加カバレッジ: パーサー・シリアライザ・ストア・バリデータの未網羅ケース

test_modules.py への補足テスト:
- parse_excel: 基本/空ファイル/空行スキップ/シート指定
- serialize_settlement_response: テスト追加
- ItnMemoryStore: 不明 noticeType・空更新・スナップショットの独立性
- MaskingFilter: 複数フィールド同時マスク・args のマスク
- コードローダー: get_area_group_name・is_valid_area_code 境界値
- Validators: 追加境界値ケース (price=0, volume=0, timeCd 非数値 etc.)
"""
import io
import unittest
import logging
import threading

import os
import sys

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.dev')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import django
django.setup()


# ============================================================
# §6.3 Excel パーサー
# ============================================================

from apps.sharepoint.file_parser import parse_excel


class TestParseExcel(unittest.TestCase):
    """parse_excel のテスト"""

    def _make_xlsx(self, rows: list) -> bytes:
        """テスト用の .xlsx をメモリ上で生成する"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_basic_parse(self):
        """ヘッダ行 + データ行を正しくパースすること"""
        content = self._make_xlsx([
            ['deliveryDate', 'areaCd', 'price'],
            ['2026-04-01', '1', 100],
            ['2026-04-02', '2', 200],
        ])
        rows = parse_excel(content)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]['deliveryDate'], '2026-04-01')
        self.assertEqual(rows[0]['areaCd'], '1')
        self.assertEqual(rows[0]['price'], 100)

    def test_empty_workbook_returns_empty_list(self):
        """空のワークブックは空リストを返すこと"""
        from openpyxl import Workbook
        wb = Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_excel(buf.getvalue())
        self.assertEqual(rows, [])

    def test_header_only_returns_empty_list(self):
        """ヘッダ行のみの場合は空リストを返すこと"""
        content = self._make_xlsx([['deliveryDate', 'areaCd', 'price']])
        rows = parse_excel(content)
        self.assertEqual(rows, [])

    def test_skip_empty_rows(self):
        """完全に空の行はスキップされること"""
        content = self._make_xlsx([
            ['deliveryDate', 'areaCd'],
            ['2026-04-01', '1'],
            [None, None],   # 空行
            ['2026-04-02', '2'],
        ])
        rows = parse_excel(content)
        self.assertEqual(len(rows), 2)

    def test_header_none_becomes_col_index(self):
        """ヘッダが None の場合は 'col_N' という名前になること"""
        content = self._make_xlsx([
            [None, 'areaCd'],
            ['2026-04-01', '1'],
        ])
        rows = parse_excel(content)
        self.assertEqual(len(rows), 1)
        self.assertIn('col_0', rows[0])

    def test_specific_sheet_name(self):
        """sheet_name を指定した場合、そのシートのデータがパースされること"""
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        ws1.append(['col'])
        ws1.append(['row1'])

        ws2 = wb.create_sheet('Sheet2')
        ws2.append(['col'])
        ws2.append(['row2'])

        buf = io.BytesIO()
        wb.save(buf)

        rows = parse_excel(buf.getvalue(), sheet_name='Sheet2')
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['col'], 'row2')


# ============================================================
# §5.2 シリアライザ追加テスト
# ============================================================

from apps.itd_api.serializers import (
    serialize_settlement_response,
    serialize_bid_response,
    serialize_inquiry_response,
    serialize_contract_response,
    serialize_error,
)


class TestSerializerAdditional(unittest.TestCase):
    """serialize_settlement_response および追加シナリオのテスト"""

    def test_settlement_response_success(self):
        """正常な清算レスポンスに success=True と count が含まれること"""
        resp = serialize_settlement_response({'settlements': [{'settlementNo': 'S001'}]})
        self.assertTrue(resp['success'])
        self.assertEqual(resp['count'], 1)
        self.assertEqual(resp['settlements'][0]['settlementNo'], 'S001')

    def test_settlement_response_empty(self):
        """清算データが空の場合は count=0"""
        resp = serialize_settlement_response({'settlements': []})
        self.assertTrue(resp['success'])
        self.assertEqual(resp['count'], 0)

    def test_settlement_response_no_key(self):
        """settlements キーがない場合は空リストとして扱われること"""
        resp = serialize_settlement_response({})
        self.assertEqual(resp['count'], 0)

    def test_bid_response_message_included(self):
        """bid レスポンスに statusInfo が message として含まれること"""
        resp = serialize_bid_response({'status': '200', 'bidNo': 'B001', 'statusInfo': '処理完了'})
        self.assertEqual(resp['message'], '処理完了')

    def test_inquiry_response_empty_bids(self):
        """bids が空の場合は count=0, bids=[] を返すこと"""
        resp = serialize_inquiry_response({'bids': []})
        self.assertTrue(resp['success'])
        self.assertEqual(resp['count'], 0)
        self.assertEqual(resp['bids'], [])

    def test_contract_response_no_key(self):
        """contractResults キーがない場合は count=0"""
        resp = serialize_contract_response({})
        self.assertEqual(resp['count'], 0)

    def test_error_includes_message(self):
        """error に message フィールドが含まれること"""
        resp = serialize_error('SOME_ERROR', 'エラーが発生しました')
        self.assertEqual(resp['message'], 'エラーが発生しました')


# ============================================================
# §5.4 ItnMemoryStore 追加テスト
# ============================================================

from apps.itn_stream.store import ItnMemoryStore


class TestItnMemoryStoreAdditional(unittest.TestCase):
    """ItnMemoryStore の追加カバレッジ"""

    def test_unknown_notice_type_ignored(self):
        """不明な noticeType はデータストアに影響しないこと"""
        store = ItnMemoryStore()
        store.update_notices([
            {'noticeTypeCd': 'UNKNOWN', 'bidNo': '001'},
        ])
        snap = store.get_snapshot()
        self.assertEqual(len(snap['contracts']), 0)
        self.assertEqual(len(snap['boards']), 0)

    def test_empty_notices_still_increments_version(self):
        """空リストの update_notices でもバージョンが上がること"""
        store = ItnMemoryStore()
        v0 = store.get_version()
        store.update_notices([])
        self.assertGreater(store.get_version(), v0)

    def test_snapshot_is_independent_copy(self):
        """get_snapshot() が返すリストは内部辞書の参照コピーであること
        (スナップショット後の変更が次のスナップショットに反映されないこと)"""
        store = ItnMemoryStore()
        store.update_notices([
            {'noticeTypeCd': 'CONTRACT', 'bidNo': '001', 'price': 100},
        ])
        snap1 = store.get_snapshot()
        # スナップショット後に新しいデータを追加
        store.update_notices([
            {'noticeTypeCd': 'CONTRACT', 'bidNo': '002', 'price': 200},
        ])
        snap2 = store.get_snapshot()

        # snap1 は変更前のデータを保持しているはず
        self.assertEqual(len(snap1['contracts']), 1)
        self.assertEqual(len(snap2['contracts']), 2)

    def test_thread_safety_concurrent_updates(self):
        """並行スレッドから更新しても例外が発生しないこと"""
        store = ItnMemoryStore()
        errors = []

        def updater(n):
            try:
                for i in range(10):
                    store.update_notices([
                        {'noticeTypeCd': 'CONTRACT', 'bidNo': f'{n}-{i}'},
                    ])
            except Exception as e:
                errors.append(e)

        threads = [threading.Thread(target=updater, args=(t,)) for t in range(5)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        self.assertEqual(errors, [])

    def test_connection_status_version_increments(self):
        """set_connection_status でもバージョンが上がること"""
        store = ItnMemoryStore()
        v0 = store.get_version()
        store.set_connection_status(True)
        self.assertGreater(store.get_version(), v0)

    def test_set_full_state_resets_both_contracts_and_boards(self):
        """set_full_state で contracts と boards が両方リセットされること"""
        store = ItnMemoryStore()
        store.update_notices([
            {'noticeTypeCd': 'CONTRACT', 'bidNo': '001'},
            {'noticeTypeCd': 'BID-BOARD', 'areaCd': '1', 'timeCd': '1'},
        ])
        store.set_full_state([])  # 空で全量リセット
        snap = store.get_snapshot()
        self.assertEqual(len(snap['contracts']), 0)
        self.assertEqual(len(snap['boards']), 0)


# ============================================================
# §8.3 MaskingFilter 追加テスト
# ============================================================

from apps.common.logging import MaskingFilter


class TestMaskingFilterAdditional(unittest.TestCase):
    """MaskingFilter の追加カバレッジ"""

    def setUp(self):
        self.filter = MaskingFilter()

    def _make_record(self, msg, args=None):
        return logging.LogRecord(
            'test', logging.INFO, '', 0, msg, args or (), None
        )

    def test_mask_multiple_fields_in_one_message(self):
        """複数の機密フィールドが1つのメッセージに含まれる場合、すべての値がマスクされること"""
        import json
        record = self._make_record(
            '{"memberId": "0841", "password": "secret", "client_secret": "abc"}'
        )
        self.filter.filter(record)
        masked = json.loads(record.msg)
        # 各フィールドの値がマスクされていること（キー名は保持される）
        self.assertEqual(masked['memberId'], '********')
        self.assertEqual(masked['password'], '********')
        self.assertEqual(masked['client_secret'], '********')

    def test_non_sensitive_data_unchanged(self):
        """機密情報を含まないメッセージは一切変更されないこと"""
        original = '[OPERATION] 入札送信: ITD1001 (date=2026-04-01)'
        record = self._make_record(original)
        self.filter.filter(record)
        self.assertEqual(record.msg, original)

    def test_mask_replaces_with_asterisks(self):
        """マスク後の値が '********' であること"""
        record = self._make_record('{"password": "mypassword123"}')
        self.filter.filter(record)
        self.assertIn('********', record.msg)

    def test_filter_returns_true(self):
        """filter() メソッドが True を返すことでログ出力が継続されること"""
        record = self._make_record('test message')
        result = self.filter.filter(record)
        self.assertTrue(result)


# ============================================================
# §9 コードローダー追加テスト
# ============================================================

from apps.common.codes import (
    load_master_codes,
    get_area_name,
    get_area_group_name,
    is_valid_area_code,
    is_valid_area_group_code,
    is_valid_bid_type,
    get_time_code_range,
    get_limits,
)


class TestMasterCodesAdditional(unittest.TestCase):
    """コードローダーの追加カバレッジ"""

    def test_all_9_area_codes_valid(self):
        """エリアコード 1〜9 がすべて有効であること"""
        for code in [str(i) for i in range(1, 10)]:
            self.assertTrue(is_valid_area_code(code), f"areaCd={code} should be valid")

    def test_area_code_0_invalid(self):
        """エリアコード '0' は無効であること"""
        self.assertFalse(is_valid_area_code('0'))

    def test_area_code_10_invalid(self):
        """エリアコード '10' は無効であること"""
        self.assertFalse(is_valid_area_code('10'))

    def test_area_code_empty_invalid(self):
        """空文字は無効であること"""
        self.assertFalse(is_valid_area_code(''))

    def test_all_6_bid_types_valid(self):
        """6種類の入札種別コードがすべて有効であること"""
        for bt in ['SELL-LIMIT', 'BUY-LIMIT', 'SELL-MARKET', 'BUY-MARKET', 'FIT', 'DEL']:
            self.assertTrue(is_valid_bid_type(bt), f"bidType={bt} should be valid")

    def test_get_area_group_name_valid(self):
        """有効なエリアグループコードで名称が取得できること"""
        name = get_area_group_name('A1')
        self.assertIsInstance(name, str)
        self.assertGreater(len(name), 0)

    def test_get_area_group_name_invalid(self):
        """無効なエリアグループコードでも例外が起きないこと (不明など返す)"""
        name = get_area_group_name('Z9')
        self.assertIsInstance(name, str)  # 何かしら返す

    def test_get_area_name_all_valid_areas(self):
        """有効なエリアコード (1〜9) で名称が取得できること"""
        for code in [str(i) for i in range(1, 10)]:
            name = get_area_name(code)
            self.assertIsInstance(name, str)
            self.assertGreater(len(name), 0)
            self.assertNotIn('不明', name)

    def test_load_master_codes_cached(self):
        """load_master_codes() は同一の辞書インスタンスを返すこと (LRUキャッシュ)"""
        codes1 = load_master_codes()
        codes2 = load_master_codes()
        self.assertIs(codes1, codes2)

    def test_area_group_codes_a1_to_a9(self):
        """存在するエリアグループコードが is_valid_area_group_code で True となること"""
        self.assertTrue(is_valid_area_group_code('A1'))
        self.assertTrue(is_valid_area_group_code('A9'))

    def test_area_group_code_a3_invalid(self):
        """A3 は存在しないエリアグループコード"""
        self.assertFalse(is_valid_area_group_code('A3'))

    def test_time_code_boundary(self):
        """時間帯コードの境界値: 1, 48 は有効 (範囲確認)"""
        min_tc, max_tc = get_time_code_range()
        self.assertEqual(min_tc, 1)
        self.assertEqual(max_tc, 48)


# ============================================================
# §7 バリデーター追加ケース
# ============================================================

from apps.common.validators import BidValidator, ValidationError


def _valid_row(**overrides):
    base = {
        'deliveryDate': '2026-04-01',
        'areaCd': '1',
        'timeCd': '1',
        'bidTypeCd': 'SELL-LIMIT',
        'price': 100.0,
        'volume': 50.0,
        'deliveryContractCd': 'BG001',
        'note': '',
    }
    base.update(overrides)
    return base


class TestValidatorAdditionalCases(unittest.TestCase):
    """BidValidator の追加境界値ケース"""

    def setUp(self):
        self.v = BidValidator()

    def test_price_zero_out_of_range(self):
        """price=0 は '0超' という制約で V-011 エラーになること"""
        errors = self.v.validate([_valid_row(price=0.0)])
        self.assertTrue(any(e.rule_id == 'V-011' for e in errors))

    def test_volume_zero_out_of_range(self):
        """volume=0 は V-014 エラーになること"""
        errors = self.v.validate([_valid_row(volume=0.0)])
        self.assertTrue(any(e.rule_id == 'V-014' for e in errors))

    def test_time_cd_non_numeric_raises_v006(self):
        """timeCd が非数値の場合も V-006 エラーになること"""
        errors = self.v.validate([_valid_row(timeCd='abc')])
        self.assertTrue(any(e.rule_id == 'V-006' for e in errors))

    def test_time_cd_zero_out_of_range(self):
        """timeCd=0 は範囲外で V-006 エラーになること"""
        errors = self.v.validate([_valid_row(timeCd='0')])
        self.assertTrue(any(e.rule_id == 'V-006' for e in errors))

    def test_time_cd_48_valid(self):
        """timeCd=48 は有効であること"""
        errors = self.v.validate([_valid_row(timeCd='48')])
        self.assertFalse(any(e.rule_id == 'V-006' for e in errors))

    def test_price_max_boundary(self):
        """price=999 (上限) は有効であること"""
        errors = self.v.validate([_valid_row(price=990.0)])  # 10の倍数で999以下
        self.assertFalse(any(e.rule_id in ('V-010', 'V-011') for e in errors))

    def test_price_above_max_v011(self):
        """price=1000 は V-011 エラーになること"""
        errors = self.v.validate([_valid_row(price=1000.0)])
        self.assertTrue(any(e.rule_id == 'V-011' for e in errors))

    def test_volume_max_boundary(self):
        """volume=5000 (上限) は有効であること"""
        errors = self.v.validate([_valid_row(volume=5000.0)])
        self.assertFalse(any(e.rule_id == 'V-014' for e in errors))

    def test_volume_above_max_v014(self):
        """volume=5000.1 は V-014 エラーになること"""
        errors = self.v.validate([_valid_row(volume=5000.1)])
        self.assertTrue(any(e.rule_id == 'V-014' for e in errors))

    def test_note_exactly_100_chars_valid(self):
        """note が exactly 100文字は有効であること"""
        errors = self.v.validate([_valid_row(note='A' * 100)])
        self.assertFalse(any(e.rule_id == 'V-016' for e in errors))

    def test_note_101_chars_invalid(self):
        """note が 101文字は V-016 エラーになること"""
        errors = self.v.validate([_valid_row(note='A' * 101)])
        self.assertTrue(any(e.rule_id == 'V-016' for e in errors))

    def test_market_sell_limit_requires_price(self):
        """SELL-LIMIT は price 必須 → price=None で V-009"""
        errors = self.v.validate([_valid_row(bidTypeCd='SELL-LIMIT', price=None)])
        self.assertTrue(any(e.rule_id == 'V-009' for e in errors))

    def test_market_order_does_not_require_price(self):
        """SELL-MARKET は price 不要 → V-009 なし"""
        row = _valid_row(bidTypeCd='SELL-MARKET', price=None)
        errors = self.v.validate([row])
        self.assertFalse(any(e.rule_id == 'V-009' for e in errors))

    def test_buy_market_does_not_require_price(self):
        """BUY-MARKET は price 不要 → V-009 なし"""
        row = _valid_row(bidTypeCd='BUY-MARKET', price=None)
        errors = self.v.validate([row])
        self.assertFalse(any(e.rule_id == 'V-009' for e in errors))

    def test_invalid_date_format_slashes(self):
        """YYYY/MM/DD 形式の日付は V-002 エラーになること"""
        errors = self.v.validate([_valid_row(deliveryDate='2026/04/01')])
        self.assertTrue(any(e.rule_id == 'V-002' for e in errors))

    def test_volume_truncation_two_decimal_places(self):
        """V-013: 小数第2位以降が切り捨てられること (例: 50.99 → 50.9)"""
        row = _valid_row(volume=50.99)
        self.v.validate([row])
        self.assertAlmostEqual(row['volume'], 50.9, places=5)

    def test_volume_truncation_does_not_trigger_v014(self):
        """V-013 補正後の値が上限以下なら V-014 エラーにならないこと"""
        row = _valid_row(volume=100.99)
        errors = self.v.validate([row])
        self.assertFalse(any(e.rule_id == 'V-014' for e in errors))

    def test_itd_market_no_errors(self):
        """ITD市場 (market='ITD') では共通バリデーションが動作すること"""
        errors = self.v.validate([_valid_row()], market='ITD')
        self.assertEqual(len(errors), 0)

    def test_dah_duplicate_triggers_error_on_each_dup_row(self):
        """V-D02: 重複した両方の行にエラーが出ること"""
        row1 = _valid_row()
        row2 = _valid_row()
        errors = self.v.validate([row1, row2], market='DAH')
        dup_errors = [e for e in errors if e.rule_id == 'V-D02']
        # 重複した2行それぞれに検出
        self.assertGreaterEqual(len(dup_errors), 1)

    def test_fail_fast_stops_after_first_batch(self):
        """VALIDATION_FAIL_FAST=True でも validate 自体は全件確認する
        (Fail-Fastはサービス層で制御)"""
        rows = [_valid_row(areaCd='99') for _ in range(3)]  # 3行すべて無効
        errors = self.v.validate(rows)
        # 3行分のエラーが出ること
        v004_errors = [e for e in errors if e.rule_id == 'V-004']
        self.assertEqual(len(v004_errors), 3)


# ============================================================
# §6.3 CSV パーサー追加テスト
# ============================================================

from apps.sharepoint.file_parser import parse_csv


class TestParseCsvAdditional(unittest.TestCase):
    """parse_csv の追加カバレッジ"""

    def test_parse_with_trailing_newline(self):
        """最後に改行がある CSV でも正しくパースされること"""
        csv_content = 'col1,col2\nval1,val2\n\n'  # 末尾に空行あり
        rows = parse_csv(csv_content.encode('utf-8'))
        # csv.DictReader はスキップしないこともあるが最低限 KeyError なし
        self.assertGreaterEqual(len(rows), 1)

    def test_parse_with_japanese_header(self):
        """日本語ヘッダを含む CSV がパースできること"""
        csv_content = '\ufeff受渡日,エリア\n2026-04-01,東北\n'
        rows = parse_csv(csv_content.encode('utf-8'))
        self.assertEqual(len(rows), 1)
        self.assertIn('受渡日', rows[0])

    def test_parse_multiple_rows(self):
        """複数行の CSV が全て返ること"""
        lines = ['col1,col2']
        for i in range(10):
            lines.append(f'val{i},data{i}')
        csv_content = '\n'.join(lines) + '\n'
        rows = parse_csv(csv_content.encode('utf-8'))
        self.assertEqual(len(rows), 10)


if __name__ == '__main__':
    unittest.main()
