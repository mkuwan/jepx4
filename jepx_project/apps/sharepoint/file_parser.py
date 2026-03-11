"""CSV/Excel ファイルパーサー (§6.3)

SharePointからダウンロードしたファイル(bytes)をパースし、
dict list として返す。
"""
import csv
import io
from openpyxl import load_workbook


def parse_csv(content: bytes) -> list[dict]:
    """SharePointから取得したCSVファイルのバイナリをBOM付きUTF-8としてデコードし、
    1行目（ヘッダー行）をキーとした辞書(dict)のリストに変換する。

    Args:
        content: CSVファイルの内容 (bytes)

    Returns:
        list of dict (1行=1dict, ヘッダをキーとして使用)
    """
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def parse_excel(content: bytes, sheet_name: str | None = None) -> list[dict]:
    """SharePointから取得したExcel(.xlsx)のバイナリを openpyxl で開き、
    指定されたシート（未指定時は最初のシート）の1行目をキーとした辞書(dict)のリストに変換する。
    ※完全に空の行は自動的にパース対象からスキップされます。

    Args:
        content: Excelファイルの内容 (bytes)
        sheet_name: 対象シート名 (Noneなら最初のシート)

    Returns:
        list of dict (1行=1dict, ヘッダをキーとして使用)
    """
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue  # 空行スキップ
        result.append({
            headers[i]: cell for i, cell in enumerate(row)
        })

    wb.close()
    return result
